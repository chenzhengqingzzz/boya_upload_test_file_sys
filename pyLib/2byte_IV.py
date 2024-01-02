import os
import re
import time
from tkinter import filedialog

from openpyxl import Workbook
import tkinter as tk

from openpyxl.chart import LineChart, Reference

'''开始时间'''
print(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
start_time = time.time()


def get_filePath_fileName_fileExt(filename):
    (filepath, tempfilename) = os.path.split(filename)
    (shotname, extension) = os.path.splitext(tempfilename)
    return filepath, shotname, extension

def get_filePath_list(raw_path,file_list = []): 
    for lists in os.listdir(raw_path): 
        path = os.path.join(raw_path, lists) 
        if os.path.isdir(path): 
            get_filePath_list(path,file_list)
        elif path.endswith('.dlg' or path.endswith('.txt')):
            file_list.append(path)
    return file_list

'''读取txt文件夹'''
root = tk.Tk()
root.withdraw()
# file_path = filedialog.askopenfilename()
file_path = filedialog.askdirectory()
file_list = get_filePath_list(file_path)

for file in file_list:
    
    filepath1, shotname1, extension1 = get_filePath_fileName_fileExt(file)
    '''遍历'''
    f = open(file, 'r', encoding='UTF-8-sig')
    lines = f.readlines()
    f.close()
    dest_filename = filepath1 + r'/' + shotname1 + r'.xlsx'

    '''建excel工作簿'''
    wb = Workbook()
    ws1 = wb.active

    '''局部变量'''

    flag1 = False
    ce_row = 1
    uA_col = 2
    '''读取文件有效数据'''
    for i in range(16):
        _ = ws1.cell(column=i + 2, row=1, value='bit' + str(i))
    for line in lines:
        searchObj = re.match(r'.*602.A.*tm_IV_input .* CE_HV = (.*) V, SI_V = .*', line)  # ([^A-Za-z_]* V)
        searchObj2 = re.match(r'.*tm_IV_input.*V/12.5 V\s*([^A-Za-z_]*)\s?([a-zA-Z_]+)/250uA.*PASS ', line)

        '''插入数据到excel'''
        if searchObj:
            ce_row += 1
            _ = ws1.cell(column=1, row=ce_row, value=searchObj.group(1)+'V')
            uA_col = 2

        if searchObj2:
            IuA = float(searchObj2.group(1))
            unit = searchObj2.group(2)
            if unit == 'nA':
                IuA /= 1000.0
            elif unit == 'A':
                pass
            _ = ws1.cell(column=uA_col, row=ce_row, value=IuA)
            uA_col += 1

    '''作图'''
    chart = LineChart()
    chart.title = '2byte_IV'
    chart.height = 17.2
    chart.width = 28.6
    chart.legend.position = 'b'

    data = Reference(ws1, min_col=2, min_row=1, max_col=17, max_row=ce_row)
    categories = Reference(ws1, min_col=1, min_row=2, max_col=1, max_row=ce_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    chart.y_axis.title = 'Current (uA)'
    ws1.add_chart(chart, 'C7')

    '''保存excel'''
    wb.save(filename=dest_filename)
    print(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
