#/usr/bin/env python3
# -*- coding: utf-8 -*-
"""导入需要的模块"""
import os
import re
import math
import time
import sys

"""定义list"""
need_file = []
addrs =[]
ces = []
dut0s = []
dut1s = []
dut2s = []
dut3s = []
vt_datas=[]
addrs_2 =[]
ces_2 = []
results=[]
result_sum={}
result_sum_2={}


def delete_repeat(one_list):

    temp_list=[]
    for one in one_list:
        if one not in temp_list:
            temp_list.append(one)
    return temp_list


def adddict2(thedict,one_list,one_dict):
    for one in one_list:
        if one in thedict:
            thedict[one]=one_dict
        else:
            thedict.update({one:one_dict})
		

def adddict4(thedict, key_a, key_b, key_c, val):
    if key_a in thedict:
        if key_b in thedict[key_a]:
            if key_c in thedict[key_a][key_b]:
                thedict[key_a][key_b][key_c]=val
            else:
                thedict[key_a][key_b].update({key_c:val})
        else:
            thedict[key_a].update({key_b:{key_c:val}})
    else:
        thedict.update({key_a:{key_b:{key_c:val}}}) 

def adddict5(thedict, key_a, key_b, key_c, key_d, val):
    if key_a in thedict:
        if key_b in thedict[key_a]:
            if key_c in thedict[key_a][key_b]:
                if key_d in thedict[key_a][key_b][key_c]:
                    thedict[key_a][key_b][key_c][key_d]=val
                else:
                    thedict[key_a][key_b][key_c].update({key_d:val})
            else:
                thedict[key_a][key_b].update({key_c:{key_d:val}})
        else:
            thedict[key_a].update({key_b:{key_c:{key_d:val}}})
    else:
        thedict.update({key_a:{key_b:{key_c:{key_d:val}}}}) 

def adddict6(thedict, key_a, key_b, key_c, key_d, key_e, val):
    if key_a in thedict:
        if key_b in thedict[key_a]:
            if key_c in thedict[key_a][key_b]:
                if key_d in thedict[key_a][key_b][key_c]:
                    if key_e in thedict[key_a][key_b][key_c][key_d]:
                        thedict[key_a][key_b][key_c][key_d][key_e]=val
                    else:
                        thedict[key_a][key_b][key_c][key_d].update({key_e:val})
                else:
                    thedict[key_a][key_b][key_c].update({key_d:{key_e:val}})
            else:
                thedict[key_a][key_b].update({key_c:{key_d:{key_e:val}}})
        else:
            thedict[key_a].update({key_b:{key_c:{key_d:{key_e:val}}}})
    else:
        thedict.update({key_a:{key_b:{key_c:{key_d:{key_e:val}}}}})


start_time = time.strftime('%Y-%m-%d %H:%M',time.localtime())
print("进程开始："+start_time)


"""文件处理"""
from openpyxl import Workbook
wb = Workbook()     # 创建工作簿
ws = wb.active      #获取动态工作表
file_path = os.getcwd()		#定义当前目录
dirs = os.listdir(file_path)        #用于返回指定的文件夹包含的文件或文件夹的名字的列表
#print("需要处理的文件列表:")
for file in dirs:
        file_name = re.search(r'(.*).txt$',file)       #捕捉需要的文件
        if file_name:
                #print(file_name.group())
                need_file_name=file_name.group()
                need_file.append(file_name.group())
                #sheet_name = file_name.group(1)
                sheet_name = "VT"
                print("正在匹配文件：{}...".format(file_name.group()))
                #ws = wb.create_sheet()
                ws.title = sheet_name

                with open(file,'r',encoding='UTF-8-sig', errors='ignore') as f:
                    for line in f:
                        searchobj = re.search(
                        #r'.* block: (0x0001), CE: (1.10V), DUT 0, cnt: 0, (), DUT 1, cnt: 0, (), DUT 2, cnt: 0, (), DUT 3, cnt: 0, () .*',line)
                        r'.*\s+(addr|sector|block): (?P<addr>[a-zA-Z0-9]+), CE: (?P<ce>[\d\.]+)V, '
                        r'DUT 0, cnt: .*, (?P<dut0>[\-\d]+), DUT 1, cnt: .*, (?P<dut1>[\-\d]+), DUT 2, cnt: .*, (?P<dut2>[\-\d]+), DUT 3, cnt: .*, (?P<dut3>[\-\d]+) .*',line)       #捕捉需要的行
                        if searchobj:
                            addrs.append(searchobj.group('addr'))
                            addr=searchobj.group('addr')
                            ces.append(searchobj.group('ce'))
                            ce=searchobj.group('ce')
                            dut0s.append(searchobj.group('dut0'))
                            dut0=searchobj.group('dut0')
                            dut1s.append(searchobj.group('dut1'))
                            dut1=searchobj.group('dut1')
                            dut2s.append(searchobj.group('dut2'))
                            dut2=searchobj.group('dut2')
                            dut3s.append(searchobj.group('dut3'))
                            dut3=searchobj.group('dut3')
                            
                            #print(searchobj.group())		#debug
                            vt_datas.append(dut0)
                            vt_datas.append(dut1)
                            vt_datas.append(dut2)
                            vt_datas.append(dut3)
                            #print(vt_datas)
                            adddict4(result_sum,need_file_name,addr,ce,vt_datas)
                            vt_datas=[]

                
"""输出数据到Excel"""

"""数据处理"""
line_num = len(addrs)		#计算匹配到的行数
print("共匹配到{}行".format(line_num))

ces_2 = delete_repeat(ces)
ces_2.sort()
addrs_2 = delete_repeat(addrs)
addrs_2.sort()
'''
print(need_file)

for c in range(0,1):
    for a in range(0,1):
        for b in range(0,1):
            #for d in range(0,1):
                file = need_file[c]
                addr = addrs_2[a]
                ce = ces_2[b]
                #vt_data=vt_datas[d]
                results =result_sum[file][addr][ce]
                print(results)
                for e in range(0,4):
                    vt=results[e]
                    print(int(vt))
'''

""" 输出表头 """
row_num =1
col_num = 3
for a in range(0,len(need_file)):
    file = need_file[a]
    ws.cell(row=row_num,column=col_num).value=file
    col_num += 4

row_num = 2
col_num = 1
for b in range(0,len(addrs_2)):
    addr = addrs_2[b]
    ws.cell(row=row_num,column=col_num).value=addr
    row_num+=1
    col_num+=1
    for a in range(0,len(ces_2)):
        ce = ces_2[a]
        ws.cell(row=row_num,column=col_num).value=ce+"V"
        row_num+=1
    col_num-=1

row_num = 2
col_num = 3
for b in range(0,len(need_file)):
    for a in range(0,4):
        ws.cell(row=row_num,column=col_num).value="dut"+str(a)
        col_num += 1


"""输出VT数据"""
row_num = 3
col_num = 3
for c in range(0,len(need_file)):
    for a in range(0,len(addrs_2)):
        for b in range(0,len(ces_2)):
            #for d in range(0,len(vt_datas)):
                file = need_file[c]
                addr = addrs_2[a]
                ce = ces_2[b]
                #vt_data=vt_datas[d]
                results =result_sum[file][addr][ce]
                #ws.cell(row=row_num,column=col_num).value=results
                for e in range(0,4):
                    vt=results[e]
                    ws.cell(row=row_num,column=col_num).value=int(vt)
                    col_num +=1
                col_num -=4
                row_num += 1
        row_num += 1
    row_num -=( len(addrs_2)*(len(ces_2)+1)  )
    col_num += 4

"""数组清零"""
addrs =[]
ces = []
dut0s = []
dut1s = []
dut2s = []
dut3s = []
addrs_2 =[]
ces_2 = []
vt_datas=[]
results=[]
result_sum = {}

"""
split() 通过指定分隔符对字符串进行切片,如果参数 num 有指定值，则分隔 num+1 个子字符串
split() 方法语法：
str.split(str="", num=string.count(str)).
str -- 分隔符，默认为所有的空字符，包括空格、换行(\n)、制表符(\t)等。
num -- 分割次数。默认为 -1, 即分隔所有。
"""

save_files = need_file[0].split(".",1)
save_file = save_files[0]
wb.save(save_file+".xlsx")			#保存
print()
print("保存为："+save_file+".xlsx")
stop_time = time.strftime('%Y-%m-%d %H:%M',time.localtime())
print("进程结束："+start_time)
print("进程耗时："+str(time.process_time())+"秒")
