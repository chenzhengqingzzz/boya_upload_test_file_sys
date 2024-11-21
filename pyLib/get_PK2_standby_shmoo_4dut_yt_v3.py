#/usr/bin/env python3
# -*- coding: utf-8 -*-
"""导入需要的模块"""
import os
import re
import math
import time
import sys
#import openpyxl
#from openpyxl.styles import Font,colors,PatternFill
#from decimal import Decimal

"""定义list"""
need_file = []
need_file2 = []
sites =[]
dpd_sites = []
vccs = []
dpd_vccs = []
vccs_2 = []
dpd_vccs_2 = []
duts = []
dpd_duts = []
standbys = []
dpds = []
result_sum={}
result_sum_2={}



def delete_repeat(one_list):

    temp_list=[]
    for one in one_list:
        if one not in temp_list:
            temp_list.append(one)
    return temp_list


def adddict2(thedict,one_list,one_dict):
    for one in one_list:
        if one in thedict:
            thedict[one]=one_dict
        else:
            thedict.update({one:one_dict})
		

def adddict4(thedict, key_a, key_b, key_c, val):
    if key_a in thedict:
        if key_b in thedict[key_a]:
            if key_c in thedict[key_a][key_b]:
                thedict[key_a][key_b][key_c]=val
            else:
                thedict[key_a][key_b].update({key_c:val})
        else:
            thedict[key_a].update({key_b:{key_c:val}})
    else:
        thedict.update({key_a:{key_b:{key_c:val}}}) 

def adddict5(thedict, key_a, key_b, key_c, key_d, val):
    if key_a in thedict:
        if key_b in thedict[key_a]:
            if key_c in thedict[key_a][key_b]:
                if key_d in thedict[key_a][key_b][key_c]:
                    thedict[key_a][key_b][key_c][key_d]=val
                else:
                    thedict[key_a][key_b][key_c].update({key_d:val})
            else:
                thedict[key_a][key_b].update({key_c:{key_d:val}})
        else:
            thedict[key_a].update({key_b:{key_c:{key_d:val}}})
    else:
        thedict.update({key_a:{key_b:{key_c:{key_d:val}}}}) 

def adddict6(thedict, key_a, key_b, key_c, key_d, key_e, val):
    if key_a in thedict:
        if key_b in thedict[key_a]:
            if key_c in thedict[key_a][key_b]:
                if key_d in thedict[key_a][key_b][key_c]:
                    if key_e in thedict[key_a][key_b][key_c][key_d]:
                        thedict[key_a][key_b][key_c][key_d][key_e]=val
                    else:
                        thedict[key_a][key_b][key_c][key_d].update({key_e:val})
                else:
                    thedict[key_a][key_b][key_c].update({key_d:{key_e:val}})
            else:
                thedict[key_a][key_b].update({key_c:{key_d:{key_e:val}}})
        else:
            thedict[key_a].update({key_b:{key_c:{key_d:{key_e:val}}}})
    else:
        thedict.update({key_a:{key_b:{key_c:{key_d:{key_e:val}}}}})

start_time = time.strftime('%Y-%m-%d %H:%M',time.localtime())
print("进程开始："+start_time)


"""文件处理"""
from openpyxl import Workbook
wb = Workbook()     # 创建工作簿
ws = wb.active      #获取动态工作表
file_path = os.getcwd()		#定义当前目录
dirs = os.listdir(file_path)        #用于返回指定的文件夹包含的文件或文件夹的名字的列表
#print("需要处理的文件列表:")
for file in dirs:
    file_name = re.search(r'(.*).dlg$',file)       #捕捉需要的文件
    if file_name:
    #print(file_name.group())
        need_file_name=file_name.group()
        need_file.append(file_name.group())
        #sheet_name = file_name.group(1)
        sheet_name = "standby_dpd_ICC"
        print("正在匹配文件：{}...".format(file_name.group()))
        #ws = wb.create_sheet()
        ws.title = sheet_name

        with open(file,'r',encoding='UTF-8-sig', errors='ignore') as f:
            for line in f:
#		    602.A00.001.***   VCC1     standby_icc_shmoo_test  1.6 V/12.5 V   1.8uA/250uA      50uA/NA         PASS 				
                searchobj = re.search(r'.*A0(?P<site>\d+).00(?P<dut>\d+).*V.*\s+standby_icc_shmoo_test\s+(?P<vcc>[\d\.]+)\s+V\/.*V\s+(?P<standby>[\d\.]+)(?P<unit>\w+A)/.*uA.*uA.*',line)       #捕捉需要的行
                searchobj_2 = re.search(r'.*A0(?P<site>\d+).00(?P<dut>\d+).*V.*\s+dpd_icc_shmoo_test\s+(?P<vcc>[\d\.]+)\s+V\/.*V\s+(?P<dpd>[\d\.]+)(?P<unit>\w+A)/.*uA.*uA.*',line)       #捕捉需要的行
#    			602.A00.003.***   VPP0     dpd_icc_shmoo_test  1.6 V/12.5 V   275nA/25uA       10uA/NA         PASS 
                if searchobj:
                    site=searchobj.group('site')
                    sites.append(searchobj.group('site'))
                    vcc=searchobj.group('vcc')
                    vccs.append(searchobj.group('vcc'))
                    dut=searchobj.group('dut')
                    duts.append(searchobj.group('dut'))
                    #standbys.append(searchobj.group('standby'))
                    
                    if searchobj.group('unit') == 'nA':
                        standby=float(searchobj.group('standby'))/1000
                        standbys.append(float(searchobj.group('standby'))/1000)
                    elif searchobj.group('unit') == 'uA':
                        standby=float(searchobj.group('standby'))
                        standbys.append(searchobj.group('standby'))
                    elif searchobj.group('unit') == 'mA':
                        standby=float(searchobj.group('standby'))*1000
                        standbys.append(float(searchobj.group('standby'))*1000)

                    adddict5(result_sum,need_file_name,site,vcc,dut,standby)
                        

                if searchobj_2:
                    dpd_site=searchobj_2.group('site')
                    dpd_sites.append(searchobj_2.group('site'))
                    dpd_vcc=searchobj_2.group('vcc')
                    dpd_vccs.append(searchobj_2.group('vcc'))
                    dpd_dut=searchobj_2.group('dut')
                    dpd_duts.append(searchobj_2.group('dut'))

                    if searchobj_2.group('unit') == 'nA':
                        dpd=float(searchobj_2.group('dpd'))/1000
                        dpds.append(float(searchobj_2.group('dpd'))/1000)
                    elif searchobj_2.group('unit') == 'uA':
                        dpd=float(searchobj_2.group('dpd'))
                        dpds.append(searchobj_2.group('dpd'))
                    elif searchobj_2.group('unit') == 'mA':
                        dpd=float(searchobj_2.group('dpd'))*1000
                        dpds.append(float(searchobj_2.group('dpd'))*1000)

                    adddict5(result_sum_2,need_file_name,dpd_site,dpd_vcc,dpd_dut,dpd)


"""输出数据到Excel"""

"""数据处理"""
line_num = len(standbys)		#计算匹配到的行数
print("共匹配到{}行".format(line_num))

vccs_2 = delete_repeat(vccs)
vccs_2.sort()
sites_2 = delete_repeat(sites)
sites_2.sort()
duts_2 = delete_repeat(duts)
duts_2.sort()

dpd_vccs_2 = delete_repeat(dpd_vccs)
dpd_vccs_2.sort()
dpd_sites_2 = delete_repeat(dpd_sites)
dpd_sites_2.sort()
dpd_duts_2 = delete_repeat(dpd_duts)
dpd_duts_2.sort()



""" 输出表头 """
row_num = 3
col_num = 1
ws.cell(row=row_num,column=col_num).value="Sandby /uA"
row_num+=1
for a in range(0,2):
    for a in range(0,len(vccs_2)):
        vcc = vccs_2[a]
        ws.cell(row=row_num,column=col_num).value=vcc+"V"
        row_num+=1
    row_num+=3

row_num = 1
col_num = 2
for b in range(0,len(need_file)):
    file = need_file[b]
    ws.cell(row=row_num,column=col_num).value=file
    row_num+=1
    for c in range(0,len(sites_2)):
        site = sites_2[c]
        ws.cell(row=row_num,column=col_num).value='site'+str(site)
        row_num+=1
        for a in range(0,len(duts_2)):
            dut=duts_2[a]
            ws.cell(row=row_num,column=col_num).value="dut"+str(dut)
            col_num += 1
        row_num -= 1
    row_num -= 1


"""输出VT数据"""

row_num = 4
col_num = 2
for c in range(0,len(need_file)):
    for a in range(0,len(sites_2)):
        for b in range(0,len(vccs_2)):
            for d in range(0,len(duts_2)):
                file = need_file[c]
                site = sites_2[a]
                vcc = vccs_2[b]
                dut=duts_2[d]
                try:
                    results=result_sum[file][site][vcc][dut]
                except:
                    results=""
                ws.cell(row=row_num,column=col_num).value=results
                col_num +=1
            col_num -=4
            row_num += 1
        row_num -= (len(vccs_2))
        col_num += 4


row_num +=((len(vccs_2))+3)
col_num -= (len(need_file)*8)
for c in range(0,len(need_file)):
    for a in range(0,len(dpd_sites_2)):
        for b in range(0,len(dpd_vccs_2)):
            for d in range(0,len(dpd_duts_2)):
                file = need_file[c]
                site = dpd_sites_2[a]
                vcc = dpd_vccs_2[b]
                dut=dpd_duts_2[d]
                try:
                    results=result_sum_2[file][site][vcc][dut]
                except:
                    results=""
                ws.cell(row=row_num,column=col_num).value=results
                col_num +=1
            col_num -=4
            row_num += 1
        row_num -= (len(dpd_vccs_2))
        col_num += 4



"""数组清零"""
addrs =[]
ces = []
dut0s = []
dut1s = []
dut2s = []
dut3s = []
addrs_2 =[]
ces_2 = []
vt_datas=[]
results=[]
result_sum = {}

"""
split() 通过指定分隔符对字符串进行切片,如果参数 num 有指定值，则分隔 num+1 个子字符串
split() 方法语法：
str.split(str="", num=string.count(str)).
str -- 分隔符，默认为所有的空字符，包括空格、换行(\n)、制表符(\t)等。
num -- 分割次数。默认为 -1, 即分隔所有。
"""

save_files = need_file[0].split(".",1)
save_file = save_files[0]
wb.save(save_file+".xlsx")			#保存
print()
print("保存为："+save_file+".xlsx")
stop_time = time.strftime('%Y-%m-%d %H:%M',time.localtime())
print("进程结束："+start_time)
print("进程耗时："+str(time.process_time())+"秒")
