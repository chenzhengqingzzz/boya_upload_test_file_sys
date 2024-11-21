# -*- coding: utf-8 -*-
"""
Created on Fri Feb  5 15:56:12 2021

@author: LZJ
"""

import os
import re
import time
from tkinter import filedialog

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, colors, Alignment
import tkinter as tk

'''开始时间'''
print(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
start_time = time.time()

'''局部变量'''
vcc_start = 1.6
vcc_inc = 0.1#0.1
vcc_end = 3.7#3.7

delay_start1 = 2
delay_inc1 = 2
delay_end1 = 498

delay_start2 = 5
delay_inc2 = 5
delay_end2 = 1000

test_num = 6

vcc_list = [[vcc_end for j in range(8)] for i in range(test_num)]
delay_time = [[0.0 for j in range(8)] for i in range(test_num)]

last_delay_list = [[0.0 for j in range(8)] for i in range(test_num)]
last_vcc_list = [[0.0 for j in range(8)] for i in range(test_num)]

inc = int((vcc_end-vcc_start)/vcc_inc) + 1 + 2
col_list = [[1+inc*i, 1+inc*i, 1+inc*i, 1+inc*i,
             1+inc*i, 1+inc*i, 1+inc*i, 1+inc*i] for i in range(test_num)]

row_list = [[4+((1+int((delay_end1-delay_start1)/delay_inc1))+4)*j for j in range(8) ] for i in range(test_num)]

row_int = [[4+((1+int((delay_end1-delay_start1)/delay_inc1))+4)*j for j in range(8) ] for i in range(test_num)]

fail_cnt = [[0 for j in range(8)] for i in range(test_num)]

'''建excel工作簿'''
wb = Workbook()
ws1 = wb.active

# 构建颜色对象，start_color表示前景色，end_color表示背景色。
green_fill = PatternFill("solid", start_color='C6EFCE')
yellow_fill = PatternFill("solid", start_color='FFFF00')
orange_fill = PatternFill('solid', start_color='FF7E00')
red_fill = PatternFill('solid', start_color='FFC7CE')
pink_fill = PatternFill('solid', start_color='99004C')
maroon_fill = PatternFill('solid', start_color='7E0023')
alignment = Alignment(horizontal='center', vertical='center')
    
def find_file():
    root = tk.Tk()
    root.withdraw()
    path = filedialog.askdirectory()
    file_list = os.listdir(path)
    new_file_list_end1 = []
    new_file_list_end2 = []
    for file in file_list:
        searchObj = re.match(r".*com_SR_all_shmoo_(.*)_(.*)V.txt", file)
        if not searchObj:
            searchObj = re.match(r".*com_reg_all_shmoo_(.*)_(.*)V.txt", file)
        if searchObj:
            if searchObj.group(1) == 's0':
                new_file_list_end1.append(path + '/' + file)
            else:
                new_file_list_end2.append(path + '/' + file)
    return new_file_list_end1,new_file_list_end2

def get_filePath_fileName_fileExt(filename):
    (filepath, tempfilename) = os.path.split(filename)
    (shotname, extension) = os.path.splitext(tempfilename)
    return filepath, shotname, extension

def get_file_data(file_path,flag):
    test_list = []
    f = open(file_path, 'r', encoding='UTF-8-sig')
    lines = f.readlines()
    f.close()
    for line in lines:
        searchObj1 = re.match(r".*Failing.*, Dut (.*) (.*) V m=(.*) (.*), (.*)!!", line)
        searchObj2 = re.match(r".*Failing.*, Dut (.*) (.*) V i=(.*) (.*), (.*)!!", line)
        searchObj3 = re.match(r".*Failing.*, Dut (.*) (.*) V (.*), (.*)!!", line)
        if searchObj1:
            searchObj = searchObj1
        elif searchObj2:
            searchObj = searchObj2
        else:
            searchObj = searchObj3
        if searchObj:
            if searchObj == searchObj3:
                test_name = str(searchObj.group(3))
            else:
                test_name = str(searchObj.group(4))
            if test_name not in test_list:
                test_list.append(test_name)
            t_num = test_list.index(test_name)
            if flag:
                dut = int(searchObj.group(1)) + 4
            else:
                dut = int(searchObj.group(1))
            last_vcc_list[t_num][dut] = vcc_list[t_num][dut]
            last_delay_list[t_num][dut] = delay_time[t_num][dut]
            vcc_list[t_num][dut] = float(searchObj.group(2))
            if searchObj == searchObj3:
                delay_time[t_num][dut] = 0
                result = searchObj.group(4)
            else:
                delay_time[t_num][dut] = float(searchObj.group(3))
                result = searchObj.group(5)
            if last_vcc_list[t_num][dut] < vcc_list[t_num][dut]:
                col_list[t_num][dut] += 1
                _ = ws1.cell(column=col_list[t_num][dut], row=row_int[t_num][dut], value=str(vcc_list
                                                                                             [t_num][dut]) + 'V')
                print(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
            elif last_vcc_list[t_num][dut] > vcc_list[t_num][dut]:
                col_list[t_num][dut] += 2
                _ = ws1.cell(column=col_list[t_num][dut], row=row_int[t_num][dut], value=str(vcc_list[t_num][dut]) + 'V')
                _ = ws1.cell(column=col_list[t_num][dut]-1, row=row_int[t_num][dut]-1, value=str(test_name))
                _ = ws1.cell(column=col_list[t_num][dut]-1, row=row_int[t_num][dut]-2, value='Dut'+str(dut))
                if t_num > 1:
                    for i in range(1+int((delay_end2-delay_start2)/delay_inc2)):
                        _ = ws1.cell(column=col_list[t_num][dut] - 1, row=row_int[t_num][dut] + i + 1, value=str(5*(i + 1)*10) + 'us')
                elif t_num:
                    for i in range(1+int((delay_end1-delay_start1)/delay_inc1)):
                        _ = ws1.cell(column=col_list[t_num][dut] - 1, row=row_int[t_num][dut] + i + 1, value=str(2*(i + 1)*10) + 'us')
                else:
                     _ = ws1.cell(column=col_list[t_num][dut] - 1, row=row_int[t_num][dut] + 1, value='RESULT')
                 
            if last_delay_list[t_num][dut] < delay_time[t_num][dut]:
                row_list[t_num][dut] += 1
            elif last_delay_list[t_num][dut] > delay_time[t_num][dut]:
                row_list[t_num][dut] = row_int[t_num][dut]
                row_list[t_num][dut] += 1
            if last_delay_list[t_num][dut] == delay_time[t_num][dut]:
                if 'PASS' in result:
                    ws1.cell(column=col_list[t_num][dut], row=row_list[t_num][dut]+1, value=result).fill = green_fill
                else:
                    fail_cnt[t_num][dut] += 1
                    ws1.cell(column=col_list[t_num][dut], row=row_list[t_num][dut]+1, value=result).fill = red_fill 
            else:
                #_ = ws1.cell(column=col_list[t_num][dut], row=row_list[t_num][dut], value=result)
                if 'PASS' in result:
                    ws1.cell(column=col_list[t_num][dut], row=row_list[t_num][dut], value=result).fill = green_fill
                else:
                    fail_cnt[t_num][dut] += 1
                    ws1.cell(column=col_list[t_num][dut], row=row_list[t_num][dut], value=result).fill = red_fill 

if __name__ == '__main__':
    file_list1,file_list2 = find_file()
    try:
        filepath1, shotname1, extension1 = get_filePath_fileName_fileExt(file_list1[0])
    except:
        filepath1, shotname1, extension1 = get_filePath_fileName_fileExt(file_list2[0])
    dest_filename = filepath1 + r'/' + shotname1.replace("_s0_1.60V",'') + r'.xlsx'
    for file in file_list1:
        get_file_data(file,0)
    for file in file_list2:
        get_file_data(file,1)
    '''保存excel'''
    wb.save(filename=dest_filename)
    print(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
    print(fail_cnt)

#[500, 603, 486, 92, 1098, 882, 184, 337]
#[481, 205, 275, 698, 875, 1157, 204, 1494]