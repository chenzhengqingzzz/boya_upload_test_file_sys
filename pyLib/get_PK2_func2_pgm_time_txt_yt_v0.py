#/usr/bin/env python3
# -*- coding: utf-8 -*-
"""导入需要的模块"""
import os
import re
import math
import time
import sys
#import openpyxl
#from openpyxl.styles import Font,colors,PatternFill
#from decimal import Decimal

"""定义list"""
need_file = []
levels =[]
addrs =[]
vccs = []
duts = []
pgmtimes = []
results = []
ddrs_2 =[]
vccs_2 = []
duts_2 =[]
result_sum={}
result_sum_2={}


def delete_repeat(one_list):

    temp_list=[]
    for one in one_list:
        if one not in temp_list:
            temp_list.append(one)
    return temp_list


def adddict2(thedict,one_list,one_dict):
    for one in one_list:
        if one in thedict:
            thedict[one]=one_dict
        else:
            thedict.update({one:one_dict})
		

def adddict4(thedict, key_a, key_b, key_c, val):
    if key_a in thedict:
        if key_b in thedict[key_a]:
            if key_c in thedict[key_a][key_b]:
                thedict[key_a][key_b][key_c]=val
            else:
                thedict[key_a][key_b].update({key_c:val})
        else:
            thedict[key_a].update({key_b:{key_c:val}})
    else:
        thedict.update({key_a:{key_b:{key_c:val}}}) 

def adddict5(thedict, key_a, key_b, key_c, key_d, val):
    if key_a in thedict:
        if key_b in thedict[key_a]:
            if key_c in thedict[key_a][key_b]:
                if key_d in thedict[key_a][key_b][key_c]:
                    thedict[key_a][key_b][key_c][key_d]=val
                else:
                    thedict[key_a][key_b][key_c].update({key_d:val})
            else:
                thedict[key_a][key_b].update({key_c:{key_d:val}})
        else:
            thedict[key_a].update({key_b:{key_c:{key_d:val}}})
    else:
        thedict.update({key_a:{key_b:{key_c:{key_d:val}}}}) 

def adddict6(thedict, key_a, key_b, key_c, key_d, key_e, val):
    if key_a in thedict:
        if key_b in thedict[key_a]:
            if key_c in thedict[key_a][key_b]:
                if key_d in thedict[key_a][key_b][key_c]:
                    if key_e in thedict[key_a][key_b][key_c][key_d]:
                        thedict[key_a][key_b][key_c][key_d][key_e]=val
                    else:
                        thedict[key_a][key_b][key_c][key_d].update({key_e:val})
                else:
                    thedict[key_a][key_b][key_c].update({key_d:{key_e:val}})
            else:
                thedict[key_a][key_b].update({key_c:{key_d:{key_e:val}}})
        else:
            thedict[key_a].update({key_b:{key_c:{key_d:{key_e:val}}}})
    else:
        thedict.update({key_a:{key_b:{key_c:{key_d:{key_e:val}}}}})

start_time = time.strftime('%Y-%m-%d %H:%M',time.localtime())
print("进程开始："+start_time)


"""文件处理"""
from openpyxl import Workbook
wb = Workbook()     # 创建工作簿
ws = wb.active      #获取动态工作表
file_path = os.getcwd()		#定义当前目录
dirs = os.listdir(file_path)        #用于返回指定的文件夹包含的文件或文件夹的名字的列表
#print("需要处理的文件列表:")
for file in dirs:
    file_name = re.search(r'(.*pgm\_time.*).txt$',file)       #捕捉需要的文件
    if file_name:
    #print(file_name.group())
        need_file.append(file_name.group())
        sheet_name = file_name.group(1)
        print("正在匹配文件：{}...".format(file_name.group()))
        ws = wb.create_sheet()
        ws.title = sheet_name

        with open(file,'r',encoding='UTF-8-sig', errors='ignore') as f:
            for line in f:
                #searchobj1 = re.search(r'.*54\[4\:3\].*\=.* (?P<level>[\d]+).*',line)	#捕捉需要的行
                searchobj1 = re.search(r'(?P<trimbit>.*\[.*\])\s+=\s+(?P<level>[\d\w\_]+).*',line)	#捕捉需要的行
                searchobj2 = re.search(r'\s+voltage\s+(?P<vcc>[\d\.]+)\s+V.*',line)	#捕捉需要的行
                #searchobj3 = re.search(r'.*period (?P<dut>[\d\.]+) us, loopcnt1 = (?P<dut>\d+).*',line)	#捕捉需要的行
                searchobj4 = re.search(r'.*dut(?P<dut>\d) page (?P<addr>[\w\W\d]+), (?P<pgmtime>[\d\.]+) us.*',line)	#捕捉需要的行
                if searchobj1:
                    trimbit=searchobj1.group('trimbit')
                    level=searchobj1.group('level')
                    levels.append(searchobj1.group('level'))
##                    adddict5(result_sum,level,vcc,dut,addr,pgmtime)
                if searchobj2:
                    vcc=searchobj2.group('vcc')
                    vccs.append(searchobj2.group('vcc'))
##                    adddict5(result_sum,level,vcc,dut,addr,pgmtime)
                if searchobj4:
                    dut=searchobj4.group('dut')
                    duts.append(searchobj4.group('dut'))
                    addr=searchobj4.group('addr')
                    addrs.append(searchobj4.group('addr'))
                    pgmtime=searchobj4.group('pgmtime')
                    pgmtimes.append(searchobj4.group('pgmtime'))

                    adddict5(result_sum,level,vcc,dut,addr,pgmtime)
##                    adddict2(result_sum_2,need_file,adddict5)
##                    print(searchobj.group())				#debug


        """输出数据到Excel"""

        """数据处理"""
        line_num = len(duts)		#计算匹配到的行数
        print("共匹配到{}行".format(line_num))

        vccs_2 = delete_repeat(vccs)
        vccs_2.sort()
        duts_2 = delete_repeat(duts)
        duts_2.sort()
        addrs_2 = delete_repeat(addrs)
        addrs_2.sort()


        print(levels)
        print(vccs_2)
        print(duts_2)
        print(addrs_2)
##        print(result_sum['00']['1.60']['0'])
##        print(result_sum)

        """ 输出表头 """
        row_num = 1
        col_num = 1
        ws.cell(row=row_num,column=col_num).value=sheet_name
        row_num = 2
        col_num = 3
        for b in range(0,len(levels)):
            level = levels[b]
            ws.cell(row=row_num,column=col_num).value= str(trimbit) +" = "+ str(level)
            col_num += (len(vccs_2))

        row_num = 3
        col_num = 1

        for a in range(0,len(duts_2)):
            dut = duts_2[a]
            ws.cell(row=row_num,column=col_num).value="dut"+str(dut)
            col_num += 2
            for b in range(0,len(levels)):
                for c in range(0,len(vccs_2)):
                    vcc = vccs_2[c]
                    ws.cell(row=row_num,column=col_num).value=vcc+"V"
                    col_num += 1
            row_num += 1
            col_num = 2
            for d in range(0,len(addrs_2)):
                addr=addrs_2[d]
                ws.cell(row=row_num,column=col_num).value=addr
                row_num +=1
            row_num += 1
            col_num -= 1

        """ 输出数据 """
        row_num = 4
        col_num = 3
        for c in range(0,len(duts_2)):
            for a in range(0,len(levels)):
                for b in range(0,len(vccs_2)):
                    for d in range(0,len(addrs_2)):
                        level = levels[a]
                        vcc = vccs_2[b]
                        dut = duts_2[c]
                        addr=addrs_2[d]
##                        print(level,vcc,dut,addr)
                        result =round( float(result_sum[level][vcc][dut][addr]), 2)
##                        print(row_num,col_num)
                        ws.cell(row=row_num,column=col_num).value=result
                        row_num +=1
                    col_num += 1
                    row_num -= (len(addrs_2))
##                col_num += 1
##                row_num -= (len(addrs_2)-1)
            row_num +=(2+len(addrs_2))
            col_num -= (len(vccs_2)*len(levels))
             


        levels = []
        duts = []
        vccs = []
        addrs = []
        pgmtimes = []
        vccs_2 = []
        duts_2 = []
        addrs_2 = []
        result_sum = {}


##save_files = in发[0].split(".",1)
save_file = need_file[0].split('.')[0]
##print(save_file)
wb.save(save_file+".xlsx")			#保存
print()
print("保存为："+file_path+"\\"+save_file+".xlsx")
stop_time = time.strftime('%Y-%m-%d %H:%M',time.localtime())
print("进程结束："+start_time)
print("进程耗时："+str(time.process_time())+"秒")
